"""
Agora Campaign helpers:
  POST /api/agora-campaigns/voice-prompt/extract-text  — extract plain text from a PDF/DOCX file
  POST /api/agora-campaigns/voice-prompt/generate      — stream AI-generated Voice Agent prompt JSON
  POST /api/agora-campaigns/quota-suggest              — AI quota cell suggestion from questionnaire
"""
from __future__ import annotations

import json
import logging

import httpx
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

import anthropic

from app.core.config import settings
from app.services.voice_prompt_generator import generate_voice_prompt_stream
from app.services.parsers.docx_parser import extract_docx_text

logger = logging.getLogger(__name__)

router = APIRouter(prefix='/api/agora-campaigns', tags=['agora-campaigns'])


def _anthropic_client() -> anthropic.AsyncAnthropic:
    return anthropic.AsyncAnthropic(
        api_key=settings.anthropic_api_key,
        http_client=httpx.AsyncClient(verify=False),
    )


# ── Text extraction ────────────────────────────────────────────────────────────

@router.post('/voice-prompt/extract-text')
async def extract_text(file: UploadFile = File(...)):
    """Extract plain text from an uploaded PDF or DOCX questionnaire file."""
    data = await file.read()
    filename = (file.filename or '').lower()

    if filename.endswith('.pdf'):
        # Return raw bytes as base64 text for PDF — caller passes it back for vision
        import base64
        return {'text': None, 'pdf_b64': base64.standard_b64encode(data).decode(), 'type': 'pdf'}

    if filename.endswith('.docx'):
        text = extract_docx_text(data)
        return {'text': text, 'type': 'docx'}

    if filename.endswith('.xlsx'):
        try:
            import openpyxl, io
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            parts: list[str] = []
            for ws in wb.worksheets:
                parts.append(f'=== {ws.title} ===')
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        parts.append('\t'.join(cells))
            return {'text': '\n'.join(parts)[:8000], 'type': 'xlsx'}
        except Exception as exc:
            logger.warning('xlsx extraction failed: %s', exc)
            return {'text': '', 'type': 'xlsx'}

    # Fallback: try UTF-8 decode
    try:
        return {'text': data.decode('utf-8', errors='replace')[:8000], 'type': 'text'}
    except Exception:
        return {'text': '', 'type': 'unknown'}


# ── Streaming prompt generation ────────────────────────────────────────────────

@router.post('/voice-prompt/generate')
async def generate_voice_prompt(
    file: UploadFile | None = File(None),
    language: str = Form('ko'),
    simplified: str = Form('false'),
):
    """Stream AI-generated Voice Agent interviewer prompt JSON."""
    if not settings.anthropic_api_key:
        raise HTTPException(503, detail='ANTHROPIC_API_KEY not configured')

    file_data: bytes | None = None
    file_type: str | None = None
    questionnaire_raw: str | None = None

    if file:
        file_data = await file.read()
        fn = (file.filename or '').lower()
        if fn.endswith('.pdf'):
            file_type = 'pdf'
        elif fn.endswith('.docx'):
            file_type = 'docx'
            questionnaire_raw = extract_docx_text(file_data)
            file_data = None
        elif fn.endswith('.xlsx'):
            try:
                import openpyxl, io
                wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
                parts: list[str] = []
                for ws in wb.worksheets:
                    parts.append(f'=== {ws.title} ===')
                    for row in ws.iter_rows(values_only=True):
                        cells = [str(c) for c in row if c is not None]
                        if cells:
                            parts.append('\t'.join(cells))
                questionnaire_raw = '\n'.join(parts)[:8000]
            except Exception:
                questionnaire_raw = ''
            file_data = None
        else:
            try:
                questionnaire_raw = file_data.decode('utf-8', errors='replace')[:8000]
            except Exception:
                questionnaire_raw = ''
            file_data = None

    is_simplified = simplified.lower() in ('true', '1', 'yes')

    async def event_stream():
        async for chunk in generate_voice_prompt_stream(
            file_data=file_data,
            file_type=file_type,
            questionnaire_raw=questionnaire_raw,
            language=language,
            simplified=is_simplified,
        ):
            yield chunk

    return StreamingResponse(event_stream(), media_type='text/plain; charset=utf-8')


# ── AI Quota Suggestion ────────────────────────────────────────────────────────

_QUOTA_SUGGEST_SYSTEM = """\
You are an expert in telephone survey quota sampling. Given a questionnaire, extract the quota/sampling constraints.

Output ONLY valid JSON in this exact format:
{
  "has_quota": true,
  "cells": [
    {"label": "string", "filters": {"DimensionName": "value"}, "target": 30}
  ],
  "message": "optional explanation when has_quota is false"
}

Rules:
- If the questionnaire has demographic quotas (age, gender, region, etc.), set has_quota=true and list each quota cell.
- Each cell's "label" should be a human-readable description (e.g. "Male, 20-29, Seoul").
- Each cell's "filters" should be key-value pairs matching the dimensions (e.g. {"Gender": "Male", "Age": "20-29", "Region": "Seoul"}).
- Set "target" to the number from the questionnaire if specified, otherwise use 30 as default.
- If no quota constraints exist, set has_quota=false and explain in "message".
- Output ONLY the JSON object, no markdown or extra text.
"""


@router.post('/quota-suggest')
async def quota_suggest(
    file: UploadFile | None = File(None),
    text: str | None = Form(None),
    language: str = Form('ko'),
):
    """Analyze questionnaire file or text and suggest quota cells using AI."""
    if not settings.anthropic_api_key:
        raise HTTPException(503, detail='ANTHROPIC_API_KEY not configured')

    questionnaire_text: str = ''

    if file:
        file_data = await file.read()
        fn = (file.filename or '').lower()
        if fn.endswith('.pdf'):
            # Use Claude's document vision for PDF
            import base64
            b64 = base64.standard_b64encode(file_data).decode()
            client = _anthropic_client()
            try:
                resp = await client.messages.create(
                    model='claude-sonnet-4-6',
                    max_tokens=4096,
                    system=_QUOTA_SUGGEST_SYSTEM,
                    messages=[{
                        'role': 'user',
                        'content': [
                            {
                                'type': 'document',
                                'source': {
                                    'type': 'base64',
                                    'media_type': 'application/pdf',
                                    'data': b64,
                                },
                            },
                            {'type': 'text', 'text': 'Analyze this questionnaire and extract quota/sampling constraints.'},
                        ],
                    }],
                )
                raw = resp.content[0].text.strip()
                if raw.startswith('```'):
                    raw = raw.split('\n', 1)[1].rsplit('```', 1)[0].strip()
                return json.loads(raw)
            except Exception as exc:
                logger.error('quota_suggest PDF error: %s', exc)
                raise HTTPException(500, detail=str(exc))
        elif fn.endswith('.docx'):
            questionnaire_text = extract_docx_text(file_data)
        elif fn.endswith('.xlsx'):
            try:
                import openpyxl, io
                wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
                parts: list[str] = []
                for ws in wb.worksheets:
                    parts.append(f'=== {ws.title} ===')
                    for row in ws.iter_rows(values_only=True):
                        cells = [str(c) for c in row if c is not None]
                        if cells:
                            parts.append('\t'.join(cells))
                questionnaire_text = '\n'.join(parts)[:8000]
            except Exception:
                questionnaire_text = ''
        else:
            try:
                questionnaire_text = file_data.decode('utf-8', errors='replace')[:8000]
            except Exception:
                questionnaire_text = ''

    if text:
        questionnaire_text = questionnaire_text + '\n' + text if questionnaire_text else text

    if not questionnaire_text.strip():
        raise HTTPException(400, detail='No questionnaire content provided')

    client = _anthropic_client()
    try:
        resp = await client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=4096,
            system=_QUOTA_SUGGEST_SYSTEM,
            messages=[{
                'role': 'user',
                'content': f'Language: {language}\n\nQuestionnaire:\n{questionnaire_text[:6000]}',
            }],
        )
        raw = resp.content[0].text.strip()
        if raw.startswith('```'):
            raw = raw.split('\n', 1)[1].rsplit('```', 1)[0].strip()
        return json.loads(raw)
    except json.JSONDecodeError as exc:
        logger.error('quota_suggest JSON parse error: %s', exc)
        raise HTTPException(500, detail='AI returned invalid JSON')
    except Exception as exc:
        logger.error('quota_suggest error: %s', exc)
        raise HTTPException(500, detail=str(exc))
